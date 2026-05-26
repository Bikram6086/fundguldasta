"""
P21: AMC portfolio holdings fetcher and parser.

Downloads monthly portfolio disclosure Excel files from AMC websites,
parses them into normalized holdings, and inserts into portfolio_holdings table.

Confirmed sources (April 2026):
  - PPFAS (Parag Parikh Flexi Cap, scheme 122639)
  - HDFC Flexi Cap (118955)
  - HDFC Mid Cap (118989)

Architecture: auto-detect column layout from header row — handles format drift gracefully.
"""

import io
import re
import logging
import calendar
import datetime
from datetime import date
from typing import Optional

import requests
import openpyxl
import psycopg2

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, */*",
}

# Scheme code → portfolio source config.
# url_template supports {month_abbr} (Apr), {month_full} (April), {year} (2026), {day} (30)
FUND_SOURCES = {
    # ── PPFAS ──────────────────────────────────────────────────────────────
    "122639": {
        "name": "Parag Parikh Flexi Cap Fund",
        "amc": "ppfas",
        "url_template": (
            "https://amc.ppfas.com/downloads/portfolio-disclosure/{year}/"
            "PPFCF_PPFAS_Monthly_Portfolio_Report_{month_full}_{day}_{year}.xlsx"
        ),
        "sheet": "PPFCF",
    },
    # ── HDFC ───────────────────────────────────────────────────────────────
    # HDFC publishes in the month AFTER disclosure (pub_year/pub_month_num)
    "118955": {
        "name": "HDFC Flexi Cap Fund",
        "amc": "hdfc",
        "url_template": (
            "https://files.hdfcfund.com/s3fs-public/{pub_year}-{pub_month_num:02d}/"
            "Monthly%20HDFC%20Flexi%20Cap%20Fund%20-%20{day}%20{month_full}%20{year}.xlsx"
        ),
    },
    "118989": {
        "name": "HDFC Mid-Cap Opportunities Fund",
        "amc": "hdfc",
        "url_template": (
            "https://files.hdfcfund.com/s3fs-public/{pub_year}-{pub_month_num:02d}/"
            "Monthly%20HDFC%20Mid%20Cap%20Fund%20-%20{day}%20{month_full}%20{year}.xlsx"
        ),
    },
    # ── Axis — consolidated all-schemes file, sheet lookup by code ─────────
    "120505": {
        "name": "Axis Midcap Fund",
        "amc": "axis",
        "url_template": (
            "https://transact.axismf.com/cms/sites/default/files/Statutory/"
            "Monthly%20Portfolio-{day}%20{month_num_2d}%20{year_2d}.xlsx"
        ),
        "sheet": "AXISMCF",
    },
    # ── Quant — consolidated all-schemes file ───────────────────────────────
    "120828": {
        "name": "Quant Small Cap Fund",
        "amc": "quant",
        "url_template": (
            "https://quantmutual.com/Admin/disclouser/Monthly_Portfolio_{day}{month_num_2d}{year}.xlsx"
        ),
        "sheet": "qSCF",
    },
    # ── SBI — consolidated all-schemes file ────────────────────────────────
    "149134": {
        "name": "SBI Balanced Advantage Fund",
        "amc": "sbi",
        "url_template": (
            "https://www.sbimf.com/docs/default-source/scheme-portfolios/"
            "all-schemes-monthly-portfolio---as-on-{day}th-{month_lower}-{year}.xlsx"
        ),
        "sheet": "SBAF",
    },
    # ── Nippon — consolidated all-schemes xlsx (despite .xls extension) ────
    "118778": {
        "name": "Nippon India Small Cap Fund",
        "amc": "nippon",
        "url_template": (
            "https://mf.nipponindiaim.com/InvestorServices/FactsheetsDocuments/"
            "NIMF-MONTHLY-PORTFOLIO-{day}-{month_full}-{year_2d}.xls"
        ),
        "sheet": "SC",
    },
    # ── Tata — consolidated all-schemes file ───────────────────────────────
    "135800": {
        "name": "Tata Digital India Fund",
        "amc": "tata",
        "url_template": (
            "https://betacms.tatamutualfund.com/system/files/{pub_year}-{pub_month_num_2d}/"
            "Monthly%20Portfolio%20as%20on%20{day}th%20{month_full}%20{year}%20%281%29.xlsx"
        ),
        "sheet": "TDIF",
    },
    # ── Motilal — consolidated all-schemes file ─────────────────────────────
    # Note: Nasdaq 100 FOF holds ETF units only (FOF), not individual stocks.
    # Overlap for this fund will be near 0 by design. Sheet included for completeness.
    "145552": {
        "name": "Motilal Oswal Nasdaq 100 FOF",
        "amc": "motilal",
        "url_template": (
            "https://www.motilaloswalmf.com/content/dam/motilal-mf/downloads/mf/"
            "month-end-portfolio/{year}/{pub_month_lower}/Motilal%20Portfolio%20{day}%20{month_full}%20{year}%20-%20Final.xlsx"
        ),
        "sheet": "YO13",
    },
    # ── Stubs — URL patterns not yet confirmed ──────────────────────────────
    "118825": {"name": "Mirae Asset Large Cap Fund",       "amc": "mirae",   "url_template": None},
    "120152": {"name": "Kotak Large Cap Fund",             "amc": "kotak",   "url_template": None},
    "118834": {"name": "Mirae Asset Large & Mid Cap Fund", "amc": "mirae",   "url_template": None},
    "119071": {"name": "DSP Midcap Fund",                  "amc": "dsp",     "url_template": None},
}

_ISIN_RE = re.compile(r'^[A-Z]{2}[A-Z0-9]{10}$')

# ---------------------------------------------------------------------------
# URL builder
# ---------------------------------------------------------------------------

def _build_url(cfg: dict, as_of: date) -> Optional[str]:
    tmpl = cfg.get("url_template")
    if not tmpl:
        return None
    month_abbrs = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    month_fulls = ["January", "February", "March", "April", "May", "June",
                   "July", "August", "September", "October", "November", "December"]
    m = as_of.month
    # Publication month = disclosure month + 1 (for AMCs that publish in the following month)
    pub_m = m % 12 + 1
    pub_y = as_of.year + (1 if m == 12 else 0)
    pub_month_lower = month_fulls[pub_m - 1].lower()
    return tmpl.format(
        year=as_of.year,
        year_2d=str(as_of.year)[-2:],
        month_num=m,
        month_num_2d=f"{m:02d}",
        month_abbr=month_abbrs[m - 1],
        month_full=month_fulls[m - 1],
        month_lower=month_fulls[m - 1].lower(),
        day=as_of.day,
        pub_year=pub_y,
        pub_month_num=pub_m,
        pub_month_num_2d=f"{pub_m:02d}",
        pub_month_lower=pub_month_lower,
    )


# ---------------------------------------------------------------------------
# Excel downloader
# ---------------------------------------------------------------------------

def _download_xlsx(url: str) -> Optional[bytes]:
    try:
        r = requests.get(url, headers=_HEADERS, timeout=30, allow_redirects=True)
        if r.status_code != 200:
            logger.warning("HTTP %s for %s", r.status_code, url)
            return None
        if len(r.content) < 5000:
            logger.warning("Response too small (%d bytes) — likely an error page", len(r.content))
            return None
        return r.content
    except Exception as exc:
        logger.error("Download failed for %s: %s", url, exc)
        return None


# ---------------------------------------------------------------------------
# Generic Excel parser — auto-detects column layout from header row
# ---------------------------------------------------------------------------

def _parse_xlsx(content: bytes, sheet_name: Optional[str] = None) -> list:
    """
    Parse an AMC portfolio Excel file.
    Returns list of dicts: {stock_isin, stock_name, sector, weight_pct,
                             market_value_lakhs, asset_type}
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    # Find header row: scan for a row containing "ISIN" as a cell value
    header_row_idx = None
    header_row = None
    for i, row in enumerate(rows):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if any("isin" in c for c in cells):
            header_row_idx = i
            header_row = [str(c).strip() if c is not None else "" for c in row]
            break

    if header_row is None:
        logger.warning("No ISIN header found in sheet '%s'", ws.title)
        return []

    h_lower = [c.lower() for c in header_row]

    def find_col(keywords):
        for kw in keywords:
            for i, h in enumerate(h_lower):
                if kw in h:
                    return i
        return None

    isin_col   = find_col(["isin"])
    name_col   = find_col(["name of", "instrument"])
    sector_col = find_col(["industry", "sector", "rating"])
    weight_col = find_col(["% to", "% nav", "net asset", "%to"])
    value_col  = find_col(["market", "fair value", "lacs", "lak"])

    if isin_col is None or weight_col is None:
        logger.warning("Could not locate ISIN or weight column. Header: %s", header_row[:10])
        return []

    # PPFAS doesn't have name in the header row (col 0 = ticker, col 1 = actual name, col 2 = ISIN)
    # Fallback: look for name in adjacent col
    if name_col is None:
        for candidate in [isin_col + 1, isin_col - 1, isin_col + 2]:
            if 0 <= candidate < len(header_row):
                name_col = candidate
                break

    holdings = []
    current_asset_type = "Equity"

    for row in rows[header_row_idx + 1:]:
        if not row or all(c is None for c in row):
            continue

        def cell(idx):
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            return str(v).strip() if v is not None else None

        raw_isin = cell(isin_col)
        if not raw_isin:
            # Section marker — detect asset type
            text = " ".join(str(c) for c in row if c is not None).lower()
            if "equity" in text and "debt" not in text:
                current_asset_type = "Equity"
            elif any(k in text for k in ["debt", "bond", "debenture", "ncd"]):
                current_asset_type = "Debt"
            elif any(k in text for k in ["money market", "cash", "t-bill"]):
                current_asset_type = "Cash"
            continue

        if not _ISIN_RE.match(raw_isin):
            continue

        name      = cell(name_col) or raw_isin
        sector    = cell(sector_col) or "Unknown"
        weight    = cell(weight_col)
        mkt_value = cell(value_col)

        try:
            weight_f = float(str(weight).replace(",", "")) if weight else None
        except ValueError:
            weight_f = None
        try:
            value_f = float(str(mkt_value).replace(",", "")) if mkt_value else None
        except ValueError:
            value_f = None

        if weight_f is None or weight_f <= 0 or weight_f > 100:
            continue

        # Clean trailing symbols from name
        name = re.sub(r'[£\*#†‡§@]+$', '', name).strip()

        holdings.append({
            "stock_isin":         raw_isin,
            "stock_name":         name[:200],
            "sector":             sector[:100],
            "weight_pct":         round(weight_f, 4),
            "market_value_lakhs": round(value_f, 2) if value_f else None,
            "asset_type":         current_asset_type,
        })

    return holdings


# ---------------------------------------------------------------------------
# DB insert
# ---------------------------------------------------------------------------

def _upsert_holdings(scheme_code: str, disclosure_date: date,
                     holdings: list, db_config: dict) -> int:
    conn = psycopg2.connect(**db_config)
    cur = conn.cursor()

    cur.execute(
        "DELETE FROM portfolio_holdings WHERE scheme_code = %s AND disclosure_date = %s",
        (scheme_code, disclosure_date)
    )

    rows_inserted = 0
    for h in holdings:
        cur.execute("""
            INSERT INTO portfolio_holdings
                (scheme_code, disclosure_date, stock_isin, stock_name,
                 weight_pct, market_value_lakhs, sector, asset_type)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            scheme_code, disclosure_date,
            h["stock_isin"], h["stock_name"],
            h["weight_pct"], h["market_value_lakhs"],
            h["sector"], h["asset_type"],
        ))
        rows_inserted += 1

    conn.commit()
    cur.close()
    conn.close()
    return rows_inserted


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def fetch_and_store(scheme_code: str, as_of: date, db_config: dict) -> dict:
    """Download, parse, and store portfolio holdings for one fund."""
    cfg = FUND_SOURCES.get(scheme_code)
    if not cfg:
        return {"scheme_code": scheme_code, "status": "unknown_fund", "holdings_count": 0}

    url = _build_url(cfg, as_of)
    if not url:
        return {"scheme_code": scheme_code, "status": "no_url_configured", "holdings_count": 0}

    logger.info("Fetching %s from %s", cfg["name"], url)
    content = _download_xlsx(url)
    if not content:
        return {"scheme_code": scheme_code, "status": "download_failed", "holdings_count": 0}

    holdings = _parse_xlsx(content, sheet_name=cfg.get("sheet"))
    if not holdings:
        return {"scheme_code": scheme_code, "status": "parse_failed", "holdings_count": 0}

    count = _upsert_holdings(scheme_code, as_of, holdings, db_config)
    logger.info("%s: stored %d holdings for %s", cfg["name"], count, as_of)
    return {
        "scheme_code":    scheme_code,
        "fund_name":      cfg["name"],
        "status":         "ok",
        "holdings_count": count,
        "disclosure_date": str(as_of),
    }


def last_month_end() -> date:
    """Return the last day of the previous month."""
    today = date.today()
    first_this = today.replace(day=1)
    last_prev = first_this - datetime.timedelta(days=1)
    return last_prev


def fetch_all_configured(as_of: Optional[date] = None, db_config: dict = None) -> list:
    """Fetch portfolio holdings for all funds that have a configured URL."""
    if as_of is None:
        as_of = last_month_end()

    results = []
    for sc, cfg in FUND_SOURCES.items():
        if cfg.get("url_template"):
            result = fetch_and_store(sc, as_of, db_config)
            results.append(result)
            print(f"  {cfg['name']}: {result['status']} ({result.get('holdings_count', 0)} holdings)")
        else:
            print(f"  {cfg['name']}: skipped (no URL configured)")
    return results


if __name__ == "__main__":
    import sys, os
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'config', '.env'))

    DB_CONFIG = {
        "host":     os.getenv("DB_HOST", "127.0.0.1"),
        "port":     int(os.getenv("DB_PORT", 5432)),
        "database": os.getenv("DB_NAME", "fundguldasta_dev"),
        "user":     os.getenv("DB_USER", "fundguldasta_user"),
        "password": os.getenv("DB_PASSWORD", ""),
    }

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

    as_of = date(2026, 4, 30)
    print(f"\nFetching portfolio holdings as of {as_of}...\n")
    results = fetch_all_configured(as_of=as_of, db_config=DB_CONFIG)

    print("\n── Summary ──")
    for r in results:
        print(f"  {r.get('fund_name', r['scheme_code'])}: "
              f"{r['status']} — {r.get('holdings_count', 0)} holdings")
