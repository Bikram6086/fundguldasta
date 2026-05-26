"""
P22: Fund manager data ingestion from AMFI Scheme Summary Documents (SSD).

Data source: AMFI portal SSD XLS files
  - MF list:     GET https://www.amfiindia.com/api/populate-mf
  - Scheme list: GET https://www.amfiindia.com/api/populate-scheme?MF_ID={id}
  - SSD file:    GET https://portal.amfiindia.com/spages/SSD_{scheme_id}.xls

Run manually when you suspect manager changes (from fundguldasta project root):
    source venv/bin/activate && python3 data/manager_ingestion.py

Or check a specific fund:
    python3 -c "from data.manager_ingestion import fetch_managers_for_scheme; print(fetch_managers_for_scheme('118989'))"
"""

import io
import re
import logging
from datetime import datetime, date
from typing import Optional

import requests
import openpyxl
import psycopg2

from config.db import get_db_config

logger = logging.getLogger(__name__)
DB_CONFIG = get_db_config()

_HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
_SESSION = None

# ---------------------------------------------------------------------------
# AMFI scheme_id mapping for our 13 verified funds (stable — won't change)
# Discovered via https://www.amfiindia.com/api/populate-mf + populate-scheme
# ---------------------------------------------------------------------------
SCHEME_ID_MAP = {
    "118825": 3178,   # Mirae Asset Large Cap Fund
    "120152": 5107,   # Kotak Large Cap Fund
    "118834": 4678,   # Mirae Asset Large & Midcap Fund
    "122639": 7086,   # Parag Parikh Flexi Cap Fund
    "118955": 858,    # HDFC Flexi Cap Fund
    "120505": 5152,   # Axis Midcap Fund
    "119071": 2004,   # DSP Midcap Fund
    "118989": 2526,   # HDFC Mid Cap Fund
    "118778": 4761,   # Nippon India Small Cap Fund
    "120828": 80,     # quant Small Cap Fund
    "149134": 12287,  # SBI Balanced Advantage Fund
    "145552": 11628,  # Motilal Oswal Nasdaq 100 FOF
    "135800": 9492,   # Tata Digital India Fund
}


def _get_session():
    global _SESSION
    if _SESSION is None:
        _SESSION = requests.Session()
        _SESSION.headers.update(_HEADERS)
    return _SESSION


def _parse_ssd_openpyxl(content: bytes) -> list[dict]:
    """Parse modern XML-based XLS via openpyxl. Returns list of manager dicts."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.worksheets[0]
    managers = []
    for row in ws.iter_rows(min_row=1, max_row=80, values_only=True):
        vals = [v for v in row if v]
        if len(vals) < 3:
            continue
        field = str(vals[1]).strip()
        value = str(vals[2]).strip()
        if "Fund Manager Name" in field:
            managers.append({"raw_name": value, "raw_since": None})
        elif "Fund Manager From Date" in field and managers:
            managers[-1]["raw_since"] = value
    return managers


def _parse_ssd_xlrd(content: bytes) -> list[dict]:
    """Parse old BIFF XLS via xlrd. Returns list of manager dicts."""
    import xlrd
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheets()[0]
    managers = []
    i = 0
    while i < ws.nrows:
        row_vals = [str(ws.cell(i, j).value).strip()
                    for j in range(ws.ncols)
                    if ws.cell(i, j).value and str(ws.cell(i, j).value).strip()]
        if len(row_vals) >= 2:
            field = row_vals[1] if len(row_vals) > 1 else ""
            value = row_vals[2] if len(row_vals) > 2 else ""
            if "Fund Manager" in field and "Name" in field and value:
                # Find corresponding From Date row (usually 2 rows later)
                since = None
                if i + 2 < ws.nrows:
                    next_row = [str(ws.cell(i + 2, j).value).strip()
                                for j in range(ws.ncols)
                                if ws.cell(i + 2, j).value]
                    if len(next_row) >= 3 and "From Date" in next_row[1]:
                        raw = next_row[2]
                        # Excel serial date → Python date
                        try:
                            serial = float(raw)
                            if serial > 1000:
                                days = int(serial) - 25569
                                since = date.fromordinal(date(1970, 1, 1).toordinal() + days)
                        except ValueError:
                            since = _parse_date_str(raw)
                managers.append({"raw_name": value, "raw_since": since})
        i += 1
    return managers


def _parse_ssd_html(content: bytes) -> list[dict]:
    """Parse HTML-format SSD (SBI uses this). Returns list of manager dicts."""
    html = content.decode("utf-8", errors="ignore")
    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', html, re.DOTALL)
    managers = []
    since_next = None
    for row in rows:
        cells = re.findall(r'<td[^>]*>(.*?)</td>', row, re.DOTALL)
        cells = [re.sub(r'<[^>]+>', '', c).strip() for c in cells if c.strip()]
        if not cells:
            continue
        joined = " ".join(cells).lower()
        if "fund manager" in joined and "fund manager type" not in joined and "fund manager from" not in joined:
            name_vals = [c for c in cells if c and "fund manager" not in c.lower() and c != ":"]
            if name_vals:
                managers.append({"raw_name": name_vals[0], "raw_since": None})
        elif "fund manager from date" in joined and managers:
            date_vals = [c for c in cells if re.search(r'\d{2}-\w{3}-\d{4}|\d{4}-\d{2}-\d{2}', c)]
            if date_vals:
                managers[-1]["raw_since"] = _parse_date_str(date_vals[0])
    return managers


def _parse_date_str(s: str) -> Optional[date]:
    """Parse various date string formats."""
    if not s:
        return None
    for fmt in ["%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d", "%B %d, %Y", "%d-%m-%Y",
                "%dth %B %Y", "%dst %B %Y", "%dnd %B %Y", "%drd %B %Y"]:
        try:
            # Strip ordinal suffixes
            clean = re.sub(r'(\d+)(st|nd|rd|th)', r'\1', s.strip())
            return datetime.strptime(clean.strip(), fmt).date()
        except ValueError:
            pass
    return None


def fetch_managers_for_scheme(scheme_code: str) -> list[dict]:
    """
    Download and parse AMFI SSD file for a scheme, returning manager records.
    Each record has: manager_name, appointment_date (or None).
    """
    ssd_id = SCHEME_ID_MAP.get(scheme_code)
    if not ssd_id:
        logger.warning("No AMFI SSD ID mapped for scheme %s", scheme_code)
        return []

    session = _get_session()
    url = f"https://portal.amfiindia.com/spages/SSD_{ssd_id}.xls"
    try:
        r = session.get(url, timeout=15)
        if r.status_code != 200 or len(r.content) < 500:
            logger.error("SSD download failed for %s: HTTP %s", scheme_code, r.status_code)
            return []

        content = r.content
        # Detect format
        if content[:2] == b'\r\n' or b'<html' in content[:50].lower():
            managers = _parse_ssd_html(content)
        elif content[:4] == b'\xd0\xcf\x11\xe0':
            managers = _parse_ssd_xlrd(content)
        else:
            managers = _parse_ssd_openpyxl(content)

        return managers
    except Exception as exc:
        logger.error("Failed to fetch/parse SSD for %s: %s", scheme_code, exc)
        return []


def upsert_managers(scheme_code: str, managers: list[dict], source: str = "AMFI-SSD") -> int:
    """
    Replace all manager records for scheme_code with new verified data.
    Returns number of rows inserted.
    """
    if not managers:
        return 0
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("DELETE FROM fund_managers WHERE scheme_code = %s", (scheme_code,))
    count = 0
    for m in managers:
        name = m.get("manager_name") or m.get("raw_name", "").strip()
        if not name or name in ("-", "NA", "N.A.", "Not applicable"):
            continue
        appt = m.get("appointment_date") or m.get("raw_since")
        cur.execute("""
            INSERT INTO fund_managers
                (scheme_code, manager_name, appointment_date, departure_date, is_current,
                 total_experience_years, source, created_at, updated_at)
            VALUES (%s, %s, %s, NULL, TRUE, NULL, %s, NOW(), NOW())
        """, (scheme_code, name, appt, source))
        count += 1
    conn.commit()
    cur.close()
    conn.close()
    return count


def report_all_managers():
    """Print current fund_managers state with computed tenure."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT scheme_code, manager_name, appointment_date, total_experience_years, source, updated_at
        FROM fund_managers
        ORDER BY scheme_code, appointment_date
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    today = date.today()
    print(f"fund_managers table — {len(rows)} records ({today})")
    print("-" * 80)
    prev = None
    for scheme_code, name, appt, exp, source, updated in rows:
        if scheme_code != prev:
            print(f"\n  {scheme_code}:")
            prev = scheme_code
        tenure = f"{(today - appt).days / 365.25:.1f}yr" if appt else "?"
        print(f"    {name} | tenure at fund: {tenure} | exp: {exp}yr | src: {source}")


def log_pipeline(status, records, error=None):
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO pipeline_log (pipeline_name, run_date, started_at, completed_at, status, records_processed, error_message)
        VALUES ('manager_ingestion', CURRENT_DATE, NOW(), NOW(), %s, %s, %s)
    """, (status, records, error))
    conn.commit()
    cur.close()
    conn.close()


def main():
    """Re-fetch and update managers for all 13 verified schemes from AMFI."""
    print("=" * 60)
    print("FUND MANAGER INGESTION — AMFI SSD")
    print(f"Started: {datetime.now()}")
    print("=" * 60)

    total = 0
    errors = []
    for scheme_code in SCHEME_ID_MAP:
        managers = fetch_managers_for_scheme(scheme_code)
        if managers:
            n = upsert_managers(scheme_code, managers)
            total += n
            print(f"  {scheme_code}: {n} managers updated")
        else:
            print(f"  {scheme_code}: fetch failed — keeping existing data")
            errors.append(scheme_code)

    print()
    report_all_managers()
    log_pipeline("success" if not errors else "partial", total,
                 f"Failed: {errors}" if errors else None)
    print(f"\nDone. {total} records updated. Errors: {errors}")


if __name__ == "__main__":
    main()
